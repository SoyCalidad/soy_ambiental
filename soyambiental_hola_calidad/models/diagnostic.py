import base64
import os
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile

from odoo import api, fields, models
from openpyxl import Workbook, load_workbook


AVAILABLE_PRIORITIES = [
    ('na', 'N/A - No aplica'),
    ('0_porcent', '0% - No documentado / No existente'),
    ('25_porcent', '25% - Aplicado / No documentado'),
    ('50_porcent', '50% - Documentado / No aplicado'),
    ('75_porcent', '75% - Aplicado y documentado'),
    ('100_porcent', '100% - Aplicado, documentado y controlado')]

FIELDS = ('diagnostic4_1_ids', 'diagnostic4_2_ids', 'diagnostic4_3_ids', 'diagnostic4_4_ids',
          'diagnostic5_1_ids', 'diagnostic5_2_ids', 'diagnostic5_3_ids',
          'diagnostic6_1_1_ids', 'diagnostic6_1_2_ids', 'diagnostic6_1_3_ids', 'diagnostic6_1_4_ids', 'diagnostic6_2_1_ids', 'diagnostic6_2_2_ids',
          'diagnostic7_1_ids', 'diagnostic7_2_ids', 'diagnostic7_3_ids', 'diagnostic7_4_1_ids', 'diagnostic7_4_2_ids', 'diagnostic7_4_3_ids', 'diagnostic7_5_1_ids', 'diagnostic7_5_2_ids', 'diagnostic7_5_3_ids',
          'diagnostic8_1_ids', 'diagnostic8_2_ids',
          'diagnostic9_1_1_ids', 'diagnostic9_1_2_ids', 'diagnostic9_2_1_ids', 'diagnostic9_2_2_ids', 'diagnostic9_3_ids',
          'diagnostic10_1_ids', 'diagnostic10_2_ids', 'diagnostic10_3_ids')


class XLSHelper(models.Model):
    """Modelo para guardar la documentación de Soy Ambiental en formato xlsx
    Esto para evitar cambios en las reglas de acceso
    """
    _name = 'sga.hola_calidad.xls_helper'
    _description = 'Ayudante binario para archivos xlsx'

    name = fields.Char('Nombre')
    datas = fields.Binary('File', readonly=True)
    datas_fname = fields.Char('Filename', readonly=True)
    date_validate = fields.Datetime(string=u'Fecha evaluación')


class Clause(models.Model):
    _name = 'sga.hola_calidad.clause'
    _description = "Claúsulas"

    question = fields.Text(string=u'Pregunta ref.', required=True)
    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    chapter = fields.Selection(
        string=u'Capítulo',
        selection=[
            ('4_context', 'Contexto de la organización'),
            ('5_leadership', 'Liderazgo'),
            ('6_planning', 'Planificación'),
            ('7_support', 'Apoyo'),
            ('8_operation', 'Operación'),
            ('9_evaluation', 'Evaluación del desempeño'),
            ('10_improvement', 'Mejora')],
        required=True,
    )


class Requirement(models.Model):
    _name = 'sga.hola_calidad.requirement'
    _description = "Requirimientos"

    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    info = fields.Text(string=u'Interpretación', store=True)
    clause_id = fields.Many2one(
        string=u'Clausula', comodel_name='sga.hola_calidad.clause', required=True)
    chapter = fields.Selection(
        string=u'Capítulo', related='clause_id.chapter', store=True)
    position_excel = fields.Char(string=u'Posición en excel')


class DiagnosticLine(models.Model):
    _name = 'sga.hola_calidad.diagnostic.line'
    _description = 'Linea de análisis ambiental'

    requirement_id = fields.Many2one(
        string=u'Requisito', comodel_name='sga.hola_calidad.requirement', )
    clause_id = fields.Text(string=u'Clausula ID', store=True)
    clause = fields.Many2one(
        string=u'Claúsulas', comodel_name='sga.hola_calidad.clause', ondelete='cascade')

    info = fields.Text(string=u'Interpretación',
                       help="here is my message", store=True)

    name = fields.Text(string=u'Nombre requirimiento', store=True)
    qualification = fields.Selection(AVAILABLE_PRIORITIES,
                                     index=True,
                                     string=u'Calificación',
                                     required=True,
                                     default='na')

    observation = fields.Text(string=u'Observaciones')
    is_page = fields.Boolean('Is a page?')
    display_type = fields.Selection([
        ('line_section', 'Section'),
        ('line_note', 'Note'),
    ], default=False, help="Technical field for UX purpose.")

    requirement_name = fields.Char(related='requirement_id.name')

    diagnostic4_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic4_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic4_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic4_4_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')

    diagnostic5_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic5_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic5_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')

    diagnostic6_1_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic6_1_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic6_1_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic6_1_4_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic6_2_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic6_2_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')

    diagnostic7_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic7_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic7_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic7_4_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic7_4_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic7_4_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic7_5_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic7_5_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic7_5_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')

    diagnostic8_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')
    diagnostic8_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')

    diagnostic9_1_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic9_1_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic9_2_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic9_2_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                         ondelete='cascade')
    diagnostic9_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                       ondelete='cascade')

    diagnostic10_1_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                        ondelete='cascade')
    diagnostic10_2_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                        ondelete='cascade')
    diagnostic10_3_id = fields.Many2one(string=u'Diagnostico', comodel_name='sga.hola_calidad.diagnostic',
                                        ondelete='cascade')


class Diagnostic(models.Model):
    _name = 'sga.hola_calidad.diagnostic'
    _description = "Diagnostico"

    name = fields.Char(string=u'Nombre', required=True,
                       default=lambda self: "Análisis ambiental")
    user_id = fields.Many2one(
        string='Responsable',
        comodel_name='res.users',
        ondelete='cascade',
        default=lambda self: self.env.user and self.env.user.id or False,
    )

    company_id = fields.Many2one(string=u'Compañia', comodel_name='res.company', required=True,
                                 domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)], default=lambda self: self.env.user.company_id.id)
    date_diagnostic = fields.Datetime(
        string=u'Fecha creación', default=fields.Datetime.now, required=True)
    date_validate = fields.Datetime(
        string=u'Fecha evaluación', related='xls_helper.date_validate')

    all_clause = fields.Many2many(
        comodel_name='sga.hola_calidad.clause', string=u'Clausulas')

    state = fields.Selection(
        string=u'Estado',
        selection=[('draft', 'Previo'),
                   ('evaluate', 'Detallado'),
                   ('validate', 'Culminado')],
        default='draft',
    )

    diagnostic4_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic4_1_id')
    diagnostic4_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic4_2_id')
    diagnostic4_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic4_3_id')
    diagnostic4_4_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic4_4_id')

    diagnostic5_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic5_1_id')
    diagnostic5_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic5_2_id')
    diagnostic5_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic5_3_id')

    diagnostic6_1_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_1_1_id')
    diagnostic6_1_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_1_2_id')
    diagnostic6_1_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_1_3_id')
    diagnostic6_1_4_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_1_4_id')
    diagnostic6_2_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_2_1_id')
    diagnostic6_2_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic6_2_2_id')

    diagnostic7_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_1_id')
    diagnostic7_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_2_id')
    diagnostic7_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_3_id')
    diagnostic7_4_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic7_4_1_id')
    diagnostic7_4_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_4_2_id')
    diagnostic7_4_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_4_3_id')
    diagnostic7_5_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_5_1_id')
    diagnostic7_5_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_5_2_id')
    diagnostic7_5_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic7_5_3_id')

    diagnostic8_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic8_1_id')
    diagnostic8_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic8_2_id')

    diagnostic9_1_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic9_1_1_id')
    diagnostic9_1_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic9_1_2_id')
    diagnostic9_2_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic9_2_1_id')
    diagnostic9_2_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                          inverse_name='diagnostic9_2_2_id')
    diagnostic9_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                        inverse_name='diagnostic9_3_id')

    diagnostic10_1_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                         inverse_name='diagnostic10_1_id')
    diagnostic10_2_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                         inverse_name='diagnostic10_2_id')
    diagnostic10_3_ids = fields.One2many(string=u'Lineas', comodel_name='sga.hola_calidad.diagnostic.line',
                                         inverse_name='diagnostic10_3_id')

    def _get_xls_helper(self):
        res = self.env['sga.hola_calidad.xls_helper'].search([])
        return res

    datas = fields.Binary('File', readonly=True)
    datas_fname = fields.Char('Filename', readonly=True)
    xls_helper = fields.Many2one(comodel_name='sga.hola_calidad.xls_helper',
                                 string='Soy Ambiental', readonly=True, default=_get_xls_helper)

    diagnostic4_ids_100 = fields.Integer(
        string=u'Total Contexto 100%', compute='_get_diagnostic')
    diagnostic4_ids_75 = fields.Integer(
        string=u'Total Contexto 75%', compute='_get_diagnostic')
    diagnostic4_ids_50 = fields.Integer(
        string=u'Total Contexto 50%', compute='_get_diagnostic')
    diagnostic4_ids_25 = fields.Integer(
        string=u'Total Contexto 25%', compute='_get_diagnostic')
    diagnostic4_ids_0 = fields.Integer(
        string=u'Total Contexto 0%', compute='_get_diagnostic')

    diagnostic5_ids_100 = fields.Integer(
        string=u'Total Liderazgo 100%', compute='_get_diagnostic')
    diagnostic5_ids_75 = fields.Integer(
        string=u'Total Liderazgo 75%', compute='_get_diagnostic')
    diagnostic5_ids_50 = fields.Integer(
        string=u'Total Liderazgo 50%', compute='_get_diagnostic')
    diagnostic5_ids_25 = fields.Integer(
        string=u'Total Liderazgo 25%', compute='_get_diagnostic')
    diagnostic5_ids_0 = fields.Integer(
        string=u'Total Liderazgo 0%', compute='_get_diagnostic')

    diagnostic6_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic6_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic6_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic6_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic6_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic7_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic7_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic7_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic7_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic7_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic8_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic8_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic8_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic8_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic8_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic9_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic9_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic9_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic9_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic9_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic10_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic10_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic10_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic10_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic10_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic4_ids_total = fields.Integer(
        string=u'Total Punto 4', compute='_get_diagnostic', store=True)
    diagnostic5_ids_total = fields.Integer(
        string=u'Total Punto 5', compute='_get_diagnostic', store=True)
    diagnostic6_ids_total = fields.Integer(
        string=u'Total Punto 6', compute='_get_diagnostic', store=True)
    diagnostic7_ids_total = fields.Integer(
        string=u'Total Punto 7', compute='_get_diagnostic', store=True)
    diagnostic8_ids_total = fields.Integer(
        string=u'Total Punto 8', compute='_get_diagnostic', store=True)
    diagnostic9_ids_total = fields.Integer(
        string=u'Total Punto 9', compute='_get_diagnostic', store=True)
    diagnostic10_ids_total = fields.Integer(
        string=u'Total Punto 10', compute='_get_diagnostic', store=True)

    def _get_diagnostic(self):
        fields_ = dir(self)
        field_suffixes = ['4', '5', '6', '7', '8', '9', '10']
        field_string = 'diagnostic{}_ids_{}'
        field_string_ = 'diagnostic{}'
        field_total = 'diagnostic{}_ids_total'

        for suffix in field_suffixes:
            diagnostic_list = filter(lambda x: x.startswith(
                field_string_.format(suffix)), FIELDS)
            res = self.get_diagnostic_values(diagnostic_list)
            setattr(self, field_string.format(suffix, '100'), res[1])
            setattr(self, field_string.format(suffix, '75'), res[2])
            setattr(self, field_string.format(suffix, '50'), res[3])
            setattr(self, field_string.format(suffix, '25'), res[4])
            setattr(self, field_string.format(suffix, '0'), res[5])
            setattr(self, field_total.format(suffix), res[6])

    def get_diagnostic_values(self, diagnostic_list):
        total_100 = total_75 = total_50 = total_25 = total_0 = total_na = 0

        for line in diagnostic_list:
            line_t = getattr(self, line)
            for line_ in line_t:
                if line_.qualification == 'na':
                    total_na += 1
                if line_.qualification == '100_porcent':
                    total_100 += 1
                if line_.qualification == '75_porcent':
                    total_75 += 1
                if line_.qualification == '50_porcent':
                    total_50 += 1
                if line_.qualification == '25_porcent':
                    total_25 += 1
                if line_.qualification == '0_porcent':
                    total_0 += 1
        total_in_partials = [total_na, total_100,
                             total_75, total_50, total_25, total_0]
        sum_total = sum(total_in_partials)
        return [total_na, total_100, total_75, total_50, total_25, total_0, sum_total]

    def evaluate_diagnostic(self):
        paths = os.path.realpath(__file__)
        dirname = os.path.dirname(os.path.dirname(paths))
        newdir = os.path.join(dirname, 'data')
        workbook = load_workbook(newdir + '/data.xlsx')

        time = datetime.now()
        self.xls_helper.write({'date_validate': datetime.now()})
        if time:
            filename = self.name + ' ' + \
                       str(time.strftime("%Y-%m-%d %H:%M %p"))
        else:
            filename = self.name

        sheets = workbook.sheetnames
        sheet = workbook[sheets[1]]
        cont = 15
        cont_relle = 0

        for diagnostics in [getattr(self, x) for x in FIELDS]:
            for diagnostic_line in diagnostics:
                p_excel = diagnostic_line.requirement_id.position_excel
                if p_excel and cont < 416:
                    number = p_excel[1:]
                    if number != str(cont):
                        i = cont
                        tmp = 1
                        tmpfinal = True
                        tmp1 = 1
                        while tmpfinal == True and cont < 268 and i < 268:
                            if i not in [17, 18, 22, 23, 31, 33, 34, 35, 45, 46, 52, 56, 57, 61, 62, 63, 64, 75, 76, 75, 85, 86,
                                         91, 92, 99, 100, 101, 109, 110, 117, 118, 120, 121, 127, 128, 133, 134, 135,
                                         144, 145, 148, 151, 152, 153, 156, 157, 161, 162, 170, 171, 172, 177, 183,
                                         184, 192, 193, 194, 195, 205, 206, 211, 212, 213, 217, 218, 225, 226, 241, 249, 250, 252, 253, 266
                                         ] and i != int(number):
                                if cont < 268 and i < 268:
                                    # comvertir a string
                                    cell1 = sheet['B' + str(i)]
                                    cell1.value = 'X'
                                    tmp = 0
                                    i = i + 1
                                    tmp1 = 0
                            else:
                                if tmp == 0:
                                    print("vista tmp y i------>", i)
                                    tmpfinal = False
                                    i = int(number)
                                    break

                            if number == str(i):  # 31
                                tmpfinal = False
                                i = int(number)
                                break
                            else:
                                if tmp1 != 0:
                                    i = i + 1
                        cont = i
                    if number == str(cont):
                        number = p_excel[1:]
                        print("number------>", number)
                        if diagnostic_line.qualification == 'na':
                            letter = 'G'
                        elif diagnostic_line.qualification == '0_porcent':
                            letter = 'B'
                        elif diagnostic_line.qualification == '25_porcent':
                            letter = 'C'
                        elif diagnostic_line.qualification == '50_porcent':
                            letter = 'D'
                        elif diagnostic_line.qualification == '75_porcent':
                            letter = 'E'
                        elif diagnostic_line.qualification == '100_porcent':
                            letter = 'F'

                        cell = sheet[letter + number]
                        print(cell)
                        cell.value = 'X'
                        cont = cont + 1
                    if diagnostic_line  .observation:
                        cell2 = sheet['H' + number]
                        cell2.value = diagnostic_line.observation

        workbook.close()

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                xls_filelike = BytesIO(f.read())

        out = base64.b64encode(xls_filelike.getvalue()).decode()

        self.xls_helper.write({'datas': out, 'datas_fname': filename})
        filename += '%2Exlsx'

        return {
            'type': 'ir.actions.act_url',
            'target': 'new',
            'url': 'web/content/?model=' + self.xls_helper._name + '&id=' + str(
                self.xls_helper.id) + '&field=datas&download=true&filename=' + filename,
        }

    def change_state_eval(self):
        """ Aqui se verfica cada vez que se añade una clasula,
        para traer los requisitos relacionadas a esa clausula.
        Cambia el estado de draft a evalate,  cuando son seleccionadas las clausulas
        """

        ids_clause_list_all = self.all_clause.ids

        clausulas_analis_prev = self.search_all_clauseId_in_actualRequierments()
        if (len(ids_clause_list_all)):
            ids_clause_list = []

            print("LISTA DE Existentes ", clausulas_analis_prev)
            print("LISTA DE ALL PREVIO ", clausulas_analis_prev)

            # TODO: CUANDO SE ELIMINA AUN NO FUNCIONA; FALTA QUE SEPA CUAND
            # SE ESTA ELIMINANDO el registro

            # Al inicio no hay requisitos seleccionados, por ende pasa directo
            if (len(clausulas_analis_prev)):
                for item in ids_clause_list_all:
                    # print ("ENTRO ", item)
                    if str(item) not in clausulas_analis_prev:
                        # print ("NO ESTA ", item)
                        ids_clause_list.append(item)
            else:
                ids_clause_list = ids_clause_list_all
            ids_clause_list = ids_clause_list_all

            self.diagnostic4_1_ids = self._default_diagnostic_line_ids_v2(
                '4.1', ids_clause_list)
            self.diagnostic4_2_ids = self._default_diagnostic_line_ids_v2(
                '4.2', ids_clause_list)
            self.diagnostic4_3_ids = self._default_diagnostic_line_ids_v2(
                '4.3', ids_clause_list)
            self.diagnostic4_4_ids = self._default_diagnostic_line_ids_v2(
                '4.4', ids_clause_list)
            # print("2 ---->",self.diagnostic4_ids.ids)

            self.diagnostic5_1_ids = self._default_diagnostic_line_ids_v2(
                '5.1', ids_clause_list)
            self.diagnostic5_2_ids = self._default_diagnostic_line_ids_v2(
                '5.2', ids_clause_list)
            self.diagnostic5_3_ids = self._default_diagnostic_line_ids_v2(
                '5.3', ids_clause_list)

            self.diagnostic6_1_1_ids = self._default_diagnostic_line_ids_v2(
                '6.1.1', ids_clause_list)
            self.diagnostic6_1_2_ids = self._default_diagnostic_line_ids_v2(
                '6.1.2', ids_clause_list)
            self.diagnostic6_1_3_ids = self._default_diagnostic_line_ids_v2(
                '6.1.3', ids_clause_list)
            self.diagnostic6_1_4_ids = self._default_diagnostic_line_ids_v2(
                '6.1.4', ids_clause_list)
            self.diagnostic6_2_1_ids = self._default_diagnostic_line_ids_v2(
                '6.2.1', ids_clause_list)
            self.diagnostic6_2_2_ids = self._default_diagnostic_line_ids_v2(
                '6.2.2', ids_clause_list)

            self.diagnostic7_1_ids = self._default_diagnostic_line_ids_v2(
                '7.1', ids_clause_list)
            self.diagnostic7_2_ids = self._default_diagnostic_line_ids_v2(
                '7.2', ids_clause_list)
            self.diagnostic7_3_ids = self._default_diagnostic_line_ids_v2(
                '7.3', ids_clause_list)
            self.diagnostic7_4_1_ids = self._default_diagnostic_line_ids_v2(
                '7.4.1', ids_clause_list)
            self.diagnostic7_4_2_ids = self._default_diagnostic_line_ids_v2(
                '7.4.2', ids_clause_list)
            self.diagnostic7_4_3_ids = self._default_diagnostic_line_ids_v2(
                '7.4.3', ids_clause_list)
            self.diagnostic7_5_1_ids = self._default_diagnostic_line_ids_v2(
                '7.5.1', ids_clause_list)
            self.diagnostic7_5_2_ids = self._default_diagnostic_line_ids_v2(
                '7.5.2', ids_clause_list)
            self.diagnostic7_5_3_ids = self._default_diagnostic_line_ids_v2(
                '7.5.3', ids_clause_list)

            self.diagnostic8_1_ids = self._default_diagnostic_line_ids_v2(
                '8.1', ids_clause_list)
            self.diagnostic8_2_ids = self._default_diagnostic_line_ids_v2(
                '8.2', ids_clause_list)

            self.diagnostic9_1_1_ids = self._default_diagnostic_line_ids_v2(
                '9.1.1', ids_clause_list)
            self.diagnostic9_1_2_ids = self._default_diagnostic_line_ids_v2(
                '9.1.2', ids_clause_list)
            self.diagnostic9_2_1_ids = self._default_diagnostic_line_ids_v2(
                '9.2.1', ids_clause_list)
            self.diagnostic9_2_2_ids = self._default_diagnostic_line_ids_v2(
                '9.2.2', ids_clause_list)
            self.diagnostic9_3_ids = self._default_diagnostic_line_ids_v2(
                '9.3', ids_clause_list)

            self.diagnostic10_1_ids = self._default_diagnostic_line_ids_v2(
                '10.1', ids_clause_list)
            self.diagnostic10_2_ids = self._default_diagnostic_line_ids_v2(
                '10.2', ids_clause_list)
            self.diagnostic10_3_ids = self._default_diagnostic_line_ids_v2(
                '10.3', ids_clause_list)

        else:
            self.diagnostic4_1_ids = False
            self.diagnostic4_2_ids = False
            self.diagnostic4_3_ids = False
            self.diagnostic4_4_ids = False

            self.diagnostic5_1_ids = False
            self.diagnostic5_2_ids = False
            self.diagnostic5_3_ids = False

            self.diagnostic6_1_1_ids = False
            self.diagnostic6_1_2_ids = False
            self.diagnostic6_1_3_ids = False
            self.diagnostic6_1_4_ids = False
            self.diagnostic6_2_1_ids = False
            self.diagnostic6_2_2_ids = False

            self.diagnostic7_1_ids = False
            self.diagnostic7_2_ids = False
            self.diagnostic7_3_ids = False
            self.diagnostic7_4_1_ids = False
            self.diagnostic7_4_2_ids = False
            self.diagnostic7_4_3_ids = False
            self.diagnostic7_5_1_ids = False
            self.diagnostic7_5_2_ids = False
            self.diagnostic7_5_3_ids = False

            self.diagnostic8_1_ids = False
            self.diagnostic8_2_ids = False

            self.diagnostic9_1_1_ids = False
            self.diagnostic9_1_2_ids = False
            self.diagnostic9_2_1_ids = False
            self.diagnostic9_2_2_ids = False
            self.diagnostic9_3_ids = False

            self.diagnostic10_1_ids = False
            self.diagnostic10_2_ids = False
            self.diagnostic10_3_ids = False


        self.state = 'evaluate'


    def search_all_clauseId_in_actualRequierments(self):
        fields = FIELDS
        """ Los padres de los requisitos actuales"""
        lista_clausulas = []
        for field in fields:
            for line in getattr(self, field):
                clau_id = line.clause_id
                if clau_id not in lista_clausulas:
                    lista_clausulas.append(clau_id)

        return lista_clausulas

    def _default_diagnostic_line_ids_v2(self, vchapter, ids_clause):
        """ Devuelve los requisitos de las clausulas, de acuerdo
        a su capitulo y id_req
        """
        requirements = self.env['sga.hola_calidad.requirement'].search(
            [('name', '=like', vchapter + '%')])

        lines = [(5, 0, 0)]
        print([x.name for x in requirements])
        for req in requirements:
            if (req.clause_id.id in ids_clause):
                data = {
                    'info': req.info,
                    'name': req.complete_name,
                    'requirement_id': req.id,
                    'clause_id': req.clause_id.id,
                    'qualification': 'na',
                }
                lines.append((0, 0, data))
        return lines

    def change_state_draft(self):
        self.state = 'draft'

    def change_state_validate(self):
        self.state = 'validate'
